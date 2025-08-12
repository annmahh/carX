import re
import os
#import sys
import geo
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font

def process_file(file):
    #base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    base_dir = os.path.dirname(os.path.abspath(__file__))
    database4 = os.path.join(base_dir, 'database', 'line4.txt')
    database5 = os.path.join(base_dir, 'database', 'line5.txt')

    create_excel()
    lines = list(file.readlines())
    row = 3
    i = 0

    while i < len(lines):
        line = lines[i]
        if 'מ"ק' in line:
            id, date, number, model, locationA, locationB = '?', '?', '?', '?', '?', '?'

            id_match = re.search(r'מ"ק:\s*(\d+)', line)
            if id_match:
                id = id_match.group(1)

                date_match = re.search(r'\d{1,2}\.\d{1,2}\.\d{4}', line)
                if date_match:
                    date = date_match.group()

                i += 1
                line = lines[i]
                number_match = re.search(r'מ"ר:(\d+)', line)
                if number_match:
                    number = number_match.group(1)
                    i += 1
                    line = lines[i]

                    model = line
                    i += 1
                    line = lines[i]

                    locationA_flag = False
                    locationA = check_city(line, database4)
                    if locationA == 'בית חולים העמק':
                        locationA = 'עפולה'
                    elif locationA == 'סוכנויות ומרכזי שירות דור':
                        locationA = 'נצרת'
                    if '?' not in locationA:
                        locationA_flag = True
                    i += 1
                    line = lines[i]

                    locationB_flag = False
                    locationB = check_city(line, database5)
                    while not locationB_flag:
                        if '?' not in locationB:
                            locationB_flag = True
                            if locationA_flag == False:
                                if geo.get_geo(line):
                                    locationA = geo.get_geo(line)
                                else:
                                    locationA = '?'
                        else:
                            i += 1
                            line = lines[i]
                            if 'מ"ק' in line:
                                i -= 1
                                break
                            locationB = check_city(line, database5)

                        if locationB == 'בית חולים העמק':
                            locationB = 'עפולה'
                        elif locationB == 'סוכנויות ומרכזי שירות דור':
                            locationB = 'נצרת'
                        if '?' in locationB:
                            locationB = '?'

                        if locationA_flag == False:
                            if geo.get_geo(line):
                                locationA = geo.get_geo(line)
                            else:
                                locationA = '?'
            #print(id, date, number, model, locationA, locationB)
            update_excel(id, date, number, model, locationA, locationB, row)
            row += 1
        i += 1

def check_city(line, file):
    with open(file, 'r', encoding='utf-8') as file:
        cities = file.readlines()
    cities = [city.strip() for city in cities]
    for city in cities:
        if city in line:
            return city
    if ',' in line:
        city = line.split(',')[0].strip()
        return '? ' + city
    elif ' ' in line:
        city = line.split(' ')[0].strip()
        return '? ' + city
    else:
        return '?'
    return '?'

def create_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.merge_cells('A1:G1')
    color = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

    sheet['A1'] = 'Table'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet['A2'].fill = color

    sheet['B2'] = 'יעד'
    sheet['B2'].fill = color

    sheet['C2'] = 'מיקום'
    sheet['C2'].fill = color

    sheet['D2'] = 'דגם רכב'
    sheet['D2'].fill = color

    sheet['E2'] = 'מס רכב'
    sheet['E2'].fill = color

    sheet['F2'] = 'תאריך'
    sheet['F2'].fill = color

    sheet['G2'] = 'מ"ק'
    sheet['G2'].fill = color

    file_name = 'car.xlsx'
    home_directory = os.path.expanduser('~')
    desktop_path = os.path.join(home_directory, 'Desktop')
    file_path = os.path.join(desktop_path, file_name)

    workbook.save(file_path)

def update_excel(id, date, number, model, locationA, locationB, row):
    file_name = 'car.xlsx'
    home_directory = os.path.expanduser('~')
    desktop_path = os.path.join(home_directory, 'Desktop')
    file_path = os.path.join(desktop_path, file_name)

    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)

    sheet = workbook.active
    color = PatternFill(start_color='D5D5D5', end_color='D5D5D5', fill_type='solid')

    sheet['B2'] = 'יעד'
    sheet['C2'] = 'מיקום'
    sheet['D2'] = 'דגם רכב'
    sheet['E2'] = 'מס רכב'
    sheet['F2'] = 'תאריך'
    sheet['G2'] = 'מ"ק'

    data = [(id, date, number, model, locationA, locationB)]

    sheet[f'A{row}'].fill = color
    sheet[f'B{row}'] = locationB
    sheet[f'C{row}'] = locationA
    sheet[f'D{row}'] = model
    sheet[f'E{row}'] = number
    sheet[f'F{row}'] = date
    sheet[f'G{row}'] = id

    workbook.save(file_path)